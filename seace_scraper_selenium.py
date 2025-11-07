# seace_scraper_selenium.py
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import pandas as pd
import os, time, re, subprocess
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ===============================
# INSTALAR CHROMIUM SI NO EXISTE
# ===============================
try:
    print("🧩 Verificando instalación de Chromium...")
    subprocess.run(["playwright", "install", "chromium"], check=True)
    print("✅ Chromium instalado correctamente o ya existente.")
except Exception as e:
    print("⚠️ No se pudo instalar Chromium automáticamente:", e)

# ===============================
# CONFIGURACIÓN GENERAL
# ===============================
LOGIN_URL = "https://prod6.seace.gob.pe/auth-proveedor/"
USUARIO = "20605280057"
CLAVE = "12345678a"
SESSION_FILE = "session.json"
OUTPUT_FILE = "contratos_palabras_clave.xlsx"

PALABRAS_CLAVE = [
    "nube", "cloud", "hosting", "alojamiento", "plataforma virtual", "plataforma",
    "aula", "aula virtual", "saas", "elearning", "e-learning", "e learning",
    "capacitacion", "capacitación", "virtual", "web", "diseño", "taller", "talleres",
    "curso", "cursos", "transformación digital", "transformacion digital", "ux", "ui",
    "consultoría", "consultoria", "digital", "video", "online", "aprendizaje",
    "educación", "educacion", "educativo", "contenido", "agilidad", "asesoría",
    "asesoria", "programacion", "programación"
]

# ===============================
# FUNCIONES AUXILIARES
# ===============================
def esperar_carga(page, timeout=120000):
    try:
        page.wait_for_selector("text=Estamos cargando la información solicitada", state="detached", timeout=timeout)
    except PWTimeout:
        print("⚠️ Página lenta, continuando...")

def aplicar_estilos_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    font_header = Font(color="FFFFFF", bold=True)
    border_all = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.border = border_all
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.border = border_all
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 70)

    wb.save(filename)
    print("🎨 Formato aplicado al Excel correctamente.")

def extraer_datos_pagina(page, palabra):
    panels = page.locator("mat-expansion-panel.panel-contratacion mat-expansion-panel-header")
    count = panels.count()
    contratos = []

    patron = re.compile(rf"(?<!\w){re.escape(palabra)}(?!\w)", flags=re.IGNORECASE)
    for i in range(count):
        p_tags = panels.nth(i).locator("p")
        textos = [p_tags.nth(j).inner_text().strip() for j in range(p_tags.count())]
        if not textos:
            continue

        descripcion_completa = " ".join(textos).lower()
        if not patron.search(descripcion_completa):
            continue

        fecha_pub = ""
        for t in textos:
            if "Fecha de publicación" in t:
                fecha_pub = t.replace("Fecha de publicación:", "").strip()

        codigo = textos[0] if len(textos) > 0 else ""
        entidad = textos[1] if len(textos) > 1 else ""
        descripcion = textos[2] if len(textos) > 2 else ""
        fechas = textos[3] if len(textos) > 3 else ""

        contratos.append({
            "Palabra clave": palabra,
            "Código": codigo,
            "Entidad": entidad,
            "Descripción": descripcion,
            "Fechas": fechas,
            "Fecha de publicación": fecha_pub
        })
    return contratos

def realizar_busqueda(page, palabra):
    buscador = page.locator('input.field-group__input.h-input-md').first
    buscador.click()
    buscador.fill("")

    for letra in palabra:
        buscador.type(letra, delay=50)
    time.sleep(0.3)

    esperar_carga(page)
    page.wait_for_timeout(3000)
    print(f"🔍 Buscando '{palabra}'...")

    try:
        label_text = page.locator("text=Contrataciones registradas").first.inner_text(timeout=4000)
        if "(" in label_text and ")" in label_text:
            cantidad = int(label_text.split("(")[1].split(")")[0])
            if cantidad == 0:
                print(f"⏩ Sin resultados para '{palabra}', pasando a la siguiente palabra...")
                return []
    except Exception:
        pass

    try:
        page.wait_for_selector("mat-expansion-panel.panel-contratacion", timeout=15000)
    except PWTimeout:
        print(f"⚠️ No se detectaron resultados visibles para '{palabra}'.")
        return []

    contratos = []
    pagina = 1
    while True:
        print(f"📄 Extrayendo página {pagina} de resultados para '{palabra}'...")
        contratos += extraer_datos_pagina(page, palabra)
        try:
            next_button = page.locator('button[aria-label="Next page"]')
            if next_button.count() == 0 or next_button.get_attribute("disabled") is not None:
                break
            next_button.click()
            esperar_carga(page)
            page.wait_for_timeout(1500)
            pagina += 1
        except Exception:
            break

    print(f"✅ Total {len(contratos)} contratos obtenidos para '{palabra}'.")
    return contratos

# ===============================
# PROCESO PRINCIPAL
# ===============================
def main():
    resultados = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-blink-features=AutomationControlled"])
        context = browser.new_context(storage_state=SESSION_FILE if os.path.exists(SESSION_FILE) else None)
        page = context.new_page()

        print("Abriendo página de login...")
        page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=60000)

        if "auth-proveedor" in page.url:
            page.wait_for_selector('input[placeholder="Ingrese su RUC"]', timeout=20000)
            page.fill('input[placeholder="Ingrese su RUC"]', USUARIO)
            page.fill('input[type="password"]', CLAVE)
            page.click('button:has-text("Acceder")')
            print("→ Login enviado.")

        try:
            page.wait_for_url("**/terminos-condiciones", timeout=30000)
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.check('input[type="checkbox"]')
            page.click('button.bg-blue-700.text-white')
        except:
            pass

        page.wait_for_url("**/cotizacion/contrataciones", timeout=60000)
        print("✅ En buscador de contratos menores.")

        page.wait_for_selector('label:has-text("Servicio")', timeout=25000)
        page.locator('label:has-text("Servicio")').first.click(force=True)
        esperar_carga(page)

        page.wait_for_selector('label:has-text("Vigente")', timeout=25000)
        page.locator('label:has-text("Vigente")').first.click(force=True)
        esperar_carga(page)

        try:
            print("📄 Cambiando 'Registros por página' a 100...")
            select_container = page.locator("text=Registros por página").locator("..").locator("mat-select")
            select_container.click()
            page.wait_for_selector("span:has-text('100')", timeout=5000)
            page.locator("span:has-text('100')").click()
            esperar_carga(page)
            page.wait_for_timeout(2000)
            print("✅ Configurado a 100 registros por página.")
        except Exception as e:
            print(f"⚠️ No se pudo cambiar registros por página: {e}")

        for palabra in PALABRAS_CLAVE:
            resultados += realizar_busqueda(page, palabra)
            page.wait_for_timeout(1500)

        if resultados:
            df = pd.DataFrame(resultados)
            df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")
            aplicar_estilos_excel(OUTPUT_FILE)
            print(f"💾 {len(resultados)} registros guardados en {OUTPUT_FILE}")
        else:
            print("⚠️ No se extrajeron resultados para ninguna palabra clave.")

        context.storage_state(path=SESSION_FILE)
        print("✅ Proceso completado correctamente.")

if __name__ == "__main__":
    main()
